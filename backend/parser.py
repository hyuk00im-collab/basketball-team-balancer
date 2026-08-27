"""xlsx 업로드 파싱 (SPEC.md §1~2) 및 결측치 대체 정책 (TEAM_BALANCE_SPEC.md §6)."""

import io
import re
import statistics
from datetime import date
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from models import Member
from skills import (
    FORWARD,
    FORWARD_SKILLS,
    GUARD,
    GUARD_SKILLS,
    SKILLS_BY_POSITION,
    level_raw,
    level_weighted,
    role_scores,
)

DATA_SHEET = "선수레벨"
HEADER_ROW = 3     # SPEC.md §1 — 컬럼명 행
EXAMPLE_ROW = 4    # 예시행: 파싱 제외
FIRST_DATA_ROW = 5

# 현재 레이아웃 (SPEC.md §2)
#   A 이름 / B 출생년도 / C 나이(수식) / D 키(cm) / E 포지션
#   F~O 가드 역량 10 / P~Y 포워드 역량 10 / Z 역량평균
LAYOUT = {
    "name": 1,
    "birth_year": 2,
    "age": 3,
    "height": 4,
    "position": 5,
    "guard_start": 6,
    "forward_start": 16,
    "level": 26,
}

# 출생년도 열이 없던 이전 레이아웃 (헤더명을 못 찾았을 때의 fallback)
LEGACY_LAYOUT = {
    "name": 1,
    "birth_year": None,
    "age": 2,
    "height": 3,
    "position": 4,
    "guard_start": 5,
    "forward_start": 15,
    "level": 25,
}

MIN_BIRTH_YEAR = 1930


def _norm(value) -> str:
    """헤더 비교용 정규화: 공백 제거."""
    return re.sub(r"\s+", "", str(value or ""))


def _to_int(value) -> Optional[int]:
    try:
        if value is None or str(value).strip() == "":
            return None
        return int(round(float(value)))
    except (TypeError, ValueError):
        return None


def _to_float(value) -> Optional[float]:
    try:
        if value is None or str(value).strip() == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def resolve_columns(header: Tuple) -> Tuple[Dict[str, Optional[int]], Dict[str, int]]:
    """헤더 행에서 열 위치를 찾는다.

    열 순서가 바뀌거나 출생년도 열이 없는 예전 파일도 헤더명으로 찾아내며,
    헤더를 못 찾은 항목만 고정 위치(LAYOUT/LEGACY_LAYOUT)로 대체한다.

    반환: (기본 열 위치, {역량항목명: 열 위치})
    """
    found: Dict[str, int] = {}
    skill_cols: Dict[str, int] = {}
    skill_lookup = {_norm(s): s for s in GUARD_SKILLS + FORWARD_SKILLS}

    for col, raw in enumerate(header, start=1):
        key = _norm(raw)
        if not key:
            continue
        if key in skill_lookup:
            skill_cols.setdefault(skill_lookup[key], col)
        elif key == "이름":
            found.setdefault("name", col)
        elif key in ("출생년도", "출생연도", "생년"):
            found.setdefault("birth_year", col)
        elif key == "나이":
            found.setdefault("age", col)
        elif key.startswith("키"):
            found.setdefault("height", col)
        elif key == "포지션":
            found.setdefault("position", col)
        elif key == "역량평균":
            found.setdefault("level", col)

    # 역량 헤더를 하나도 못 찾으면 고정 위치 블록으로 대체.
    # 포지션 열 위치로 신/구 레이아웃을 구분한다.
    if not skill_cols:
        fallback = LEGACY_LAYOUT if found.get("position") == LEGACY_LAYOUT["position"] else LAYOUT
        for i, item in enumerate(GUARD_SKILLS):
            skill_cols[item] = fallback["guard_start"] + i
        for i, item in enumerate(FORWARD_SKILLS):
            skill_cols[item] = fallback["forward_start"] + i

    columns: Dict[str, Optional[int]] = {}
    for key in ("name", "age", "height", "position", "level"):
        columns[key] = found.get(key, LAYOUT[key])
    # 출생년도는 헤더가 있을 때만 사용한다 (예전 파일의 B열은 '나이'이므로 추정 금지)
    columns["birth_year"] = found.get("birth_year")
    return columns, skill_cols


def _cell(row: Tuple, col: Optional[int]):
    if not col or col > len(row):
        return None
    return row[col - 1]


def parse_workbook(data: bytes) -> Tuple[List[Member], List[str]]:
    """업로드된 xlsx 바이트를 인원 목록으로 변환한다.

    반환: (인원 목록, 경고 메시지 목록)
    """
    warnings: List[str] = []
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)

    if DATA_SHEET in wb.sheetnames:
        ws = wb[DATA_SHEET]
    else:
        ws = wb[wb.sheetnames[0]]
        warnings.append(
            f"'{DATA_SHEET}' 시트를 찾지 못해 '{ws.title}' 시트를 사용했습니다."
        )

    max_col = max(ws.max_column or 0, LAYOUT["level"])
    header_rows = list(
        ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, max_col=max_col, values_only=True)
    )
    columns, skill_cols = resolve_columns(header_rows[0] if header_rows else ())
    if columns["birth_year"] is None:
        warnings.append("출생년도 열이 없어 '나이' 열 값을 그대로 사용했습니다.")

    this_year = date.today().year
    members: List[Member] = []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=FIRST_DATA_ROW, max_col=max_col, values_only=True),
        start=FIRST_DATA_ROW,
    ):
        name = _cell(row, columns["name"])
        if name is None or str(name).strip() == "":
            break  # SPEC.md §1 — 빈 이름 행 = 데이터 끝

        name = str(name).strip()
        position = str(_cell(row, columns["position"]) or "").strip()
        if position not in (GUARD, FORWARD):
            warnings.append(f"{row_idx}행 '{name}': 포지션이 '{position}'이라 건너뜁니다.")
            continue

        items = SKILLS_BY_POSITION[position]
        skills: Dict[str, float] = {}
        for item in items:
            val = _to_float(_cell(row, skill_cols.get(item)))
            if val is not None:
                skills[item] = val

        estimated = len(skills) < len(items)
        if estimated and skills:
            warnings.append(f"{row_idx}행 '{name}': 역량 항목 일부가 비어 있어 중앙값으로 대체합니다.")

        # 나이는 '=YEAR(TODAY())-출생년도' 수식이라 캐시값이 없을 수 있으므로
        # 출생년도가 있으면 항상 현재 연도 기준으로 다시 계산한다.
        birth_year = _to_int(_cell(row, columns["birth_year"]))
        if birth_year is not None and not (MIN_BIRTH_YEAR <= birth_year <= this_year):
            warnings.append(f"{row_idx}행 '{name}': 출생년도 {birth_year}는 범위를 벗어나 무시합니다.")
            birth_year = None
        age = this_year - birth_year if birth_year is not None else _to_int(_cell(row, columns["age"]))

        sheet_level = _to_float(_cell(row, columns["level"]))
        level = sheet_level if sheet_level is not None else level_raw(position, skills)

        members.append(
            Member(
                id=f"row_{row_idx}",
                name=name,
                birth_year=birth_year,
                age=age,
                height_cm=_to_int(_cell(row, columns["height"])),
                position=position,
                skills=skills,
                level=round(level, 3),
                level_weighted=level_weighted(position, skills),
                roles=role_scores(position, skills),
                is_guest=False,
                estimated=estimated,
            )
        )

    wb.close()
    if not members:
        warnings.append("읽어들인 인원이 없습니다. 시트 형식(3행 헤더 / 5행부터 데이터)을 확인해 주세요.")
    return members, warnings


def position_medians(members: List[Member]) -> Dict[str, Dict[str, float]]:
    """포지션별·항목별 중앙값 (역량이 실제 입력된 인원만 사용)."""
    out: Dict[str, Dict[str, float]] = {}
    for position, items in SKILLS_BY_POSITION.items():
        pool = [m for m in members if m.position == position and not m.estimated and m.skills]
        table: Dict[str, float] = {}
        for item in items:
            vals = [m.skills[item] for m in pool if m.skills.get(item) is not None]
            if vals:
                table[item] = round(statistics.median(vals), 3)
        out[position] = table
    return out


def impute(members: List[Member]) -> List[Member]:
    """TEAM_BALANCE_SPEC.md §6 — 미입력 역량을 동일 포지션 중앙값으로 대체한다.

    중앙값을 낼 표본이 없으면 3.0(중앙 점수)으로 채운다.
    """
    medians = position_medians(members)
    filled: List[Member] = []
    for m in members:
        items = SKILLS_BY_POSITION[m.position]
        missing = [i for i in items if m.skills.get(i) is None]
        if not missing:
            filled.append(m)
            continue
        skills = dict(m.skills)
        for item in missing:
            skills[item] = medians[m.position].get(item, 3.0)
        copy = m.model_copy(
            update={
                "skills": skills,
                "level": level_raw(m.position, skills),
                "level_weighted": level_weighted(m.position, skills),
                "roles": role_scores(m.position, skills),
                "estimated": True,
            }
        )
        filled.append(copy)
    return filled
