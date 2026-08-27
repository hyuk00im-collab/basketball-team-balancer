"""SPEC.md 형식의 샘플/템플릿 엑셀(basketball_level_sheet.xlsx) 생성기.

웹 화면의 '샘플 파일 다운로드'(GET /api/sample)와 CLI(tools/make_sample_xlsx.py)에서
같은 코드를 사용한다.

시트에 걸어두는 자동화:
  - 나이(C)      : 출생년도(B) 기준 수식 → 파일을 여는 시점의 연도로 갱신
  - 역량평균(Z)  : 포지션(E)에 따라 가드 블록(F~O) 또는 포워드 블록(P~Y)만 평균
  - 조건부 서식  : 포지션에 해당하지 않는 역량 블록을 회색으로 가림
  - 유효성 검사  : 포지션 드롭다운, 해당 포지션 블록에만 0~5(0.5 단위) 입력 허용
"""

import random
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from parser import LAYOUT
from skills import FORWARD, FORWARD_SKILLS, GUARD, GUARD_SKILLS

FILENAME = "basketball_level_sheet.xlsx"

NAMES = [
    "김민준", "이서준", "박도윤", "최시우", "정하준", "강주원", "조지호", "윤은우",
    "장선우", "임유준", "한지훈", "오건우", "서현우", "신도현", "권민재", "황준서",
    "안예준", "송지환", "홍성민", "문태윤", "배승현", "백진우", "허재민", "남기훈",
]

EXAMPLE_ROW = 4      # 예시행 (파싱 제외)
FIRST_DATA_ROW = 5
TEMPLATE_ROWS = 100  # 수식·서식·유효성 검사를 미리 깔아두는 행 수

COL_NAME = get_column_letter(LAYOUT["name"])            # A
COL_BIRTH = get_column_letter(LAYOUT["birth_year"])     # B
COL_POSITION = get_column_letter(LAYOUT["position"])    # E
COL_G_FIRST = get_column_letter(LAYOUT["guard_start"])       # F
COL_G_LAST = get_column_letter(LAYOUT["guard_start"] + 9)    # O
COL_F_FIRST = get_column_letter(LAYOUT["forward_start"])     # P
COL_F_LAST = get_column_letter(LAYOUT["forward_start"] + 9)  # Y

AUTO_FILL = PatternFill("solid", fgColor="FFEDEFF3")    # 자동계산 열
GUARD_FILL = PatternFill("solid", fgColor="FFEAF0FF")   # 가드 역량 헤더
FORWARD_FILL = PatternFill("solid", fgColor="FFE7F6F0")  # 포워드 역량 헤더
# 조건부 서식(dxf)의 채우기는 fgColor가 아니라 bgColor를 읽는다 (ARGB로 명시)
MASK_FILL = PatternFill(bgColor="FFF2F3F5")   # 해당 없는 블록 가림
MASK_FONT = Font(color="FFEDEEF1")            # 배경과 거의 같은 색 = 글자가 안 보임


def age_formula(row: int) -> str:
    """현재 연도 기준 나이 자동 계산 (파일을 열 때마다 갱신)."""
    cell = f"{COL_BIRTH}{row}"
    return f'=IF({cell}="","",YEAR(TODAY())-{cell})'


def level_formula(row: int) -> str:
    """포지션에 따라 가드 블록 / 포워드 블록 중 하나만 평균낸다."""
    guard = f"{COL_G_FIRST}{row}:{COL_G_LAST}{row}"
    forward = f"{COL_F_FIRST}{row}:{COL_F_LAST}{row}"
    pos = f"${COL_POSITION}{row}"
    return (
        f'=IF(${COL_NAME}{row}="","",'
        f'IF({pos}="{GUARD}",IF(COUNT({guard})=0,"",AVERAGE({guard})),'
        f'IF({pos}="{FORWARD}",IF(COUNT({forward})=0,"",AVERAGE({forward})),"")))'
    )


def _add_row_formulas(ws, row: int) -> None:
    ws.cell(row=row, column=LAYOUT["age"], value=age_formula(row)).fill = AUTO_FILL
    ws.cell(row=row, column=LAYOUT["level"], value=level_formula(row)).fill = AUTO_FILL


def _add_rules(ws, last_row: int) -> None:
    """조건부 서식 + 데이터 유효성 검사 (포지션별 역량 블록 분리)."""
    guard_range = f"{COL_G_FIRST}{FIRST_DATA_ROW}:{COL_G_LAST}{last_row}"
    forward_range = f"{COL_F_FIRST}{FIRST_DATA_ROW}:{COL_F_LAST}{last_row}"
    pos = f"${COL_POSITION}{FIRST_DATA_ROW}"  # 열 고정 + 행 상대참조

    # 조건부 서식은 예시행(4행)부터 걸어 동작을 눈으로 확인할 수 있게 한다.
    # 포지션이 해당 값이 아니면 그 블록을 가린다 (포지션 미입력 행은 양쪽 모두 가림).
    cf_pos = f"${COL_POSITION}{EXAMPLE_ROW}"
    for position, first_col, last_col in (
        (GUARD, COL_G_FIRST, COL_G_LAST),
        (FORWARD, COL_F_FIRST, COL_F_LAST),
    ):
        ws.conditional_formatting.add(
            f"{first_col}{EXAMPLE_ROW}:{last_col}{last_row}",
            FormulaRule(
                formula=[f'{cf_pos}<>"{position}"'],
                fill=MASK_FILL,
                font=MASK_FONT,
                stopIfTrue=True,
            ),
        )

    # 포지션: 드롭다운
    dv_pos = DataValidation(
        type="list",
        formula1=f'"{GUARD},{FORWARD}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="포지션 확인",
        error=f"'{GUARD}' 또는 '{FORWARD}'만 선택할 수 있습니다.",
    )
    ws.add_data_validation(dv_pos)
    dv_pos.add(f"{COL_POSITION}{FIRST_DATA_ROW}:{COL_POSITION}{last_row}")

    # 역량: 해당 포지션 블록에만 0~5(0.5 단위) 입력 허용
    for position, rng, first_col in (
        (GUARD, guard_range, COL_G_FIRST),
        (FORWARD, forward_range, COL_F_FIRST),
    ):
        cell = f"{first_col}{FIRST_DATA_ROW}"
        dv = DataValidation(
            type="custom",
            formula1=f'=AND({pos}="{position}",{cell}>=0,{cell}<=5,MOD({cell}*2,1)=0)',
            showErrorMessage=True,
            errorTitle=f"{position} 역량 항목",
            error=(
                f"포지션이 '{position}'인 행에서만 입력할 수 있습니다.\n"
                "포지션을 먼저 선택하고 0~5 사이 0.5 단위로 입력하세요."
            ),
        )
        ws.add_data_validation(dv)
        dv.add(rng)


def build_workbook(rows: int = len(NAMES), seed: int = 7) -> Workbook:
    """샘플 데이터가 채워진 워크북을 만든다. rows=0이면 빈 양식만 생성."""
    wb = Workbook()

    guide = wb.active
    guide.title = "작성안내"
    guide["A1"] = "이 시트는 사람이 읽는 안내용이며 프로그램 파싱 대상이 아닙니다."
    guide["A3"] = "1) '선수레벨' 시트 3행이 헤더, 4행은 예시행(파싱 제외), 5행부터 실제 데이터입니다."
    guide["A4"] = "2) 포지션(E열)은 드롭다운에서 '가드' 또는 '포워드'를 고릅니다."
    guide["A5"] = "3) 역량 점수는 0~5, 0.5 단위입니다."
    guide["A7"] = "※ 포지션을 고르면 해당 포지션의 역량 블록만 활성화됩니다."
    guide["A8"] = "   - 가드  → F~O열(가드 역량 10항목)만 입력, P~Y열은 회색으로 가려집니다."
    guide["A9"] = "   - 포워드 → P~Y열(포워드 역량 10항목)만 입력, F~O열은 회색으로 가려집니다."
    guide["A10"] = "   - 가려진 블록에 값을 넣으려 하면 입력이 거부됩니다."
    guide["A12"] = "※ 회색 배경 열(C 나이, Z 역량평균)은 수식이므로 직접 입력하지 마세요."
    guide["A13"] = '   나이   =IF(B5="","",YEAR(TODAY())-B5)  → 파일을 여는 시점의 연도 기준'
    guide["A14"] = "   역량평균 = 포지션이 가드면 F~O 평균, 포워드면 P~Y 평균 (자동 분기)"
    guide["A16"] = f"※ 수식·서식은 {FIRST_DATA_ROW + TEMPLATE_ROWS - 1}행까지 미리 깔려 있습니다."
    guide["A17"] = "   그 아래로 더 필요하면 마지막 행을 복사해 붙여넣으세요."
    for key in ("A7", "A12", "A16"):
        guide[key].font = Font(bold=True)
    guide.column_dimensions["A"].width = 82

    ws = wb.create_sheet("선수레벨")
    ws["A1"] = "농구 동호회 선수 레벨 시트"
    ws["A2"] = "3행: 헤더 / 4행: 예시행(파싱 제외) / 5행부터 실제 데이터"

    header = (
        ["이름", "출생년도", "나이", "키(cm)", "포지션"]
        + GUARD_SKILLS
        + FORWARD_SKILLS
        + ["역량평균"]
    )
    for col, title in enumerate(header, start=1):
        cell = ws.cell(row=3, column=col, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for col in (LAYOUT["age"], LAYOUT["level"]):
        ws.cell(row=3, column=col).fill = AUTO_FILL  # 자동계산 열
    for i in range(10):  # 역량 블록 헤더 색 구분
        ws.cell(row=3, column=LAYOUT["guard_start"] + i).fill = GUARD_FILL
        ws.cell(row=3, column=LAYOUT["forward_start"] + i).fill = FORWARD_FILL
    ws.row_dimensions[3].height = 44

    this_year = date.today().year
    ws.cell(row=4, column=LAYOUT["name"], value="예시)홍길동")
    ws.cell(row=4, column=LAYOUT["birth_year"], value=this_year - 32)
    ws.cell(row=4, column=LAYOUT["height"], value=178)
    ws.cell(row=4, column=LAYOUT["position"], value=GUARD)
    for i in range(10):
        ws.cell(row=4, column=LAYOUT["guard_start"] + i, value=3.5)
    _add_row_formulas(ws, 4)

    rng = random.Random(seed)
    for offset in range(rows):
        row = FIRST_DATA_ROW + offset
        name = NAMES[offset % len(NAMES)]
        position = GUARD if offset % 2 == 0 else FORWARD
        base = rng.choice([2.0, 2.5, 3.0, 3.5, 4.0])
        ws.cell(row=row, column=LAYOUT["name"], value=name)
        ws.cell(row=row, column=LAYOUT["birth_year"], value=this_year - rng.randint(22, 48))
        ws.cell(row=row, column=LAYOUT["height"], value=rng.randint(168, 192))
        ws.cell(row=row, column=LAYOUT["position"], value=position)

        start = LAYOUT["guard_start"] if position == GUARD else LAYOUT["forward_start"]
        for i in range(10):
            score = min(5.0, max(0.0, base + rng.choice([-1.0, -0.5, 0, 0, 0.5, 1.0])))
            ws.cell(row=row, column=start + i, value=score)

    # 데이터가 없는 행에도 수식을 미리 깔아 둔다 (이름이 비면 "" 를 돌려줌)
    last_row = FIRST_DATA_ROW + max(rows, TEMPLATE_ROWS) - 1
    for row in range(FIRST_DATA_ROW, last_row + 1):
        _add_row_formulas(ws, row)
    _add_rules(ws, last_row)

    ws.column_dimensions[COL_NAME].width = 12
    ws.column_dimensions[COL_BIRTH].width = 10
    ws.column_dimensions[COL_POSITION].width = 10
    for col in range(LAYOUT["guard_start"], LAYOUT["forward_start"] + 10):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.column_dimensions[get_column_letter(LAYOUT["level"])].width = 10
    ws.freeze_panes = f"{COL_G_FIRST}4"  # 이름~포지션 열과 헤더 고정
    return wb


def save(path: Path, rows: int = len(NAMES)) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(rows).save(path)
    return path
