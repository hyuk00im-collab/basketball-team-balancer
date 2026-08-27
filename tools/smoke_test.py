"""API 스모크 테스트: 업로드 → 게스트 추가 → 3팀 편성 → 지표 확인 → xlsx 내보내기."""

import io
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "backend"))

from fastapi.testclient import TestClient  # noqa: E402

from main import app  # noqa: E402

SAMPLE = ROOT / "sample" / "basketball_level_sheet.xlsx"


def main() -> int:
    client = TestClient(app)

    # 0) 샘플/양식 다운로드 → 그대로 다시 업로드되는지(왕복) 확인
    for label, url, expect_rows in (("sample", "/api/sample", 24), ("template", "/api/sample?empty=true", 0)):
        res = client.get(url)
        assert res.status_code == 200, res.text
        assert "attachment;" in res.headers["content-disposition"], res.headers
        from openpyxl import load_workbook as _lw

        wb = _lw(io.BytesIO(res.content))
        assert wb.sheetnames == ["작성안내", "선수레벨"], wb.sheetnames
        ws = wb["선수레벨"]
        assert [ws.cell(3, c).value for c in range(1, 6)] == ["이름", "출생년도", "나이", "키(cm)", "포지션"]
        assert str(ws.cell(5, 3).value or "").startswith("=IF(B5=") or expect_rows == 0
        parsed = client.post(
            "/api/upload",
            files={"file": ("s.xlsx", res.content,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        ).json()
        assert len(parsed["added"]) == expect_rows, (label, len(parsed["added"]))
        client.delete("/api/members")
        print(f"[download] {label}: {len(res.content)} bytes → 재업로드 {expect_rows}명 OK")

    # 1) 엑셀 업로드
    with open(SAMPLE, "rb") as fp:
        res = client.post(
            "/api/upload",
            files={"file": ("basketball_level_sheet.xlsx", fp.read(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert res.status_code == 200, res.text
    uploaded = res.json()
    print(f"[upload] {len(uploaded['added'])} members, warnings={len(uploaded['warnings'])}")
    assert len(uploaded["added"]) == 24
    assert all(0 < m["level"] <= 5 for m in uploaded["added"])

    # 1-1) 출생년도 → 현재 연도 기준 나이 자동 계산 (엑셀 C열 수식과 동일 결과)
    this_year = date.today().year
    first = uploaded["added"][0]
    print(f"[age] {first['name']} birth_year={first['birth_year']} age={first['age']} (기준 {this_year})")
    assert all(m["birth_year"] for m in uploaded["added"]), "출생년도가 비어 있음"
    assert all(m["age"] == this_year - m["birth_year"] for m in uploaded["added"]), "나이 자동계산 불일치"

    # 2) 역량 없는 게스트 추가 → 중앙값 대체 확인
    res = client.post(
        "/api/members",
        json={"name": "게스트A", "position": "가드", "birth_year": this_year - 30,
              "height_cm": 180, "is_guest": True, "skills": {}},
    )
    assert res.status_code == 201, res.text
    guest = res.json()
    print(f"[guest] level={guest['level']} estimated={guest['estimated']} skills={len(guest['skills'])} age={guest['age']}")
    assert guest["estimated"] is True and len(guest["skills"]) == 10 and guest["level"] > 0
    assert guest["age"] == 30, "게스트 나이가 출생년도로 계산되지 않음"

    # 3) 역량 일부만 입력한 게스트
    res = client.post(
        "/api/members",
        json={"name": "게스트B", "position": "포워드", "skills": {"리바운드": 4.5, "블록슛": 4.0}},
    )
    assert res.status_code == 201, res.text

    # 3-1) 55세 이상 게스트 → 보너스 1점 (실질 전력 +1)
    res = client.post(
        "/api/members",
        json={"name": "노장가드", "position": "가드", "birth_year": this_year - 57, "height_cm": 176},
    )
    senior = res.json()
    print(f"[bonus] {senior['name']} age={senior['age']} level={senior['level']} "
          f"bonus={senior['age_bonus']} effective={senior['effective_level']}")
    assert senior["age_bonus"] == 0.5, "55세 이상 가중치(0.5) 미적용"
    assert abs(senior["effective_level"] - (senior["level"] + 0.5)) < 1e-9

    res = client.post(
        "/api/members",
        json={"name": "54세가드", "position": "가드", "birth_year": this_year - 54},
    )
    assert res.json()["age_bonus"] == 0.0, "54세에 보너스가 붙음"

    # 4) 3팀 편성
    res = client.post(
        "/api/teams/generate",
        json={"options": {"team_count": 3, "seed": 42, "iterations": 4000}},
    )
    assert res.status_code == 200, res.text
    result = res.json()
    sizes = [len(t["members"]) for t in result["teams"]]
    print(f"[generate] sizes={sizes} seed={result['seed']}")
    assert sum(sizes) == 28 and max(sizes) - min(sizes) <= 1

    m = result["metrics"]
    print(
        f"[metrics] 실질전력편차율(평균)={m['level_deviation_rate_avg']}% "
        f"편차율(합계)={m['level_deviation_rate']}% 쿼터위반={m['quota_violations']} "
        f"역할군지수={m['role_balance_index']} 신장편차={m['height_gap']} 창출형부재={m['creation_shortage_teams']}"
    )
    for t in result["teams"]:
        s = t["summary"]
        print(
            f"  - {t['name']}: {len(t['members'])}명 평균 {s['avg_level']} "
            f"(가드 {s['guards']}/포워드 {s['forwards']}, 평균키 {s['avg_height']}, 창출형 {s['creators']})"
        )
    assert m["quota_violations"] == 0, "포지션 쿼터 위반 발생"
    assert m["top_separation_violations"] == 0, "포지션별 상위 3명이 같은 팀에 배정됨"
    assert m["bonus_players"] == 1, "55세 이상 인원 집계 오류"
    assert m["creation_shortage_teams"] == 0, "창출형 인원 없는 팀 발생"
    assert m["level_deviation_rate_avg"] <= 5.0, "평균 레벨 편차율 5% 초과"

    # 4-1) 1순위(역량)·2순위(포워드 신장) 상위 3명이 서로 다른 팀인지 (여러 시드 반복)
    for seed in range(12):
        r = client.post("/api/teams/generate", json={"options": {"seed": seed}}).json()
        by_level, by_height = {}, {}
        for team_idx, t in enumerate(r["teams"]):
            for x in t["members"]:
                if x["position_rank"] and x["position_rank"] <= 3:
                    by_level.setdefault(x["position"], {})[x["position_rank"]] = team_idx
                if x["position"] == "포워드" and x["height_rank"] and x["height_rank"] <= 3:
                    by_height[x["height_rank"]] = team_idx
        for position, ranks in by_level.items():
            assert len(set(ranks.values())) == len(ranks), f"seed={seed} {position} 역량 상위권 중복: {ranks}"
        assert len(set(by_height.values())) == len(by_height), f"seed={seed} 포워드 키 상위권 중복: {by_height}"
        assert r["metrics"]["top_separation_violations"] == 0
        assert r["metrics"]["height_separation_violations"] == 0
    print("[top3] 12개 시드 모두 역량 1~3위 · 포워드 키 1~3위가 각각 다른 팀 OK")

    # 4-1b) 우선순위 확인 — 둘 다 만족할 수 없게 강제하면 2순위(신장)를 양보해야 한다
    everyone = client.get("/api/members").json()
    fw = [x for x in everyone if x["position"] == "포워드"]
    top_level_fw = sorted(fw, key=lambda x: (-x["level"], x["name"]))[0]
    tall = [x for x in sorted(fw, key=lambda x: (-(x["height_cm"] or 0), x["name"]))[:3]
            if x["id"] != top_level_fw["id"]][:2]
    conflict = client.post(
        "/api/teams/generate",
        json={"options": {"seed": 4, "locked": {tall[0]["id"]: 0, tall[1]["id"]: 0}}},
    ).json()
    cm = conflict["metrics"]
    print(f"[priority] 포워드 키 상위 2명을 한 팀에 고정 → 역량위반 {cm['top_separation_violations']}건 / "
          f"신장위반 {cm['height_separation_violations']}건")
    assert cm["height_separation_violations"] >= 1, "신장 위반이 보고되지 않음"
    assert cm["top_separation_violations"] == 0, "1순위(역량) 제약이 깨짐"

    # 4-2) 상위권 인원을 같은 팀에 고정하면 위반이 보고되는지
    tops = {}
    for t in result["teams"]:
        for x in t["members"]:
            if x["position"] == "가드" and x["position_rank"] and x["position_rank"] <= 2:
                tops[x["position_rank"]] = x["id"]
    forced = client.post(
        "/api/teams/generate",
        json={"options": {"seed": 3, "locked": {tops[1]: 0, tops[2]: 0}}},
    ).json()
    assert forced["metrics"]["top_separation_violations"] >= 1, "고정 배정 위반이 잡히지 않음"
    print(f"[top3] 상위 2명을 같은 팀에 고정 → 위반 {forced['metrics']['top_separation_violations']}건 보고 OK")

    # 5) 동일 시드 재현성
    res2 = client.post("/api/teams/generate", json={"options": {"seed": 42, "iterations": 4000}})
    names_a = [[x["name"] for x in t["members"]] for t in result["teams"]]
    names_b = [[x["name"] for x in t["members"]] for t in res2.json()["teams"]]
    assert names_a == names_b, "동일 시드인데 결과가 다름"
    print("[seed] 동일 시드 재현성 OK")

    # 6) 제외 인원 반영
    victim = result["teams"][0]["members"][0]["id"]
    res = client.post(
        "/api/teams/generate", json={"options": {"seed": 1, "excluded_ids": [victim]}}
    )
    excluded = res.json()
    assert sum(len(t["members"]) for t in excluded["teams"]) == 27
    assert len(excluded["excluded"]) == 1
    print("[exclude] 제외 인원 처리 OK")

    # 7) 고정 배정
    fixed_id = result["teams"][2]["members"][0]["id"]
    res = client.post(
        "/api/teams/generate", json={"options": {"seed": 5, "locked": {fixed_id: 0}}}
    )
    locked_result = res.json()
    assert any(x["id"] == fixed_id for x in locked_result["teams"][0]["members"]), "고정 배정 실패"
    print("[lock] 고정 배정 OK")

    # 7-1) 결과 화면 수동 조정 (드래그로 팀 이동) → 요약·지표 재계산
    base = client.post("/api/teams/generate", json={"options": {"seed": 42}}).json()
    ace = next(
        x for t in base["teams"] for x in t["members"]
        if x["position"] == "가드" and x["position_rank"] == 1
    )
    ace_team = next(i for i, t in enumerate(base["teams"]) if any(x["id"] == ace["id"] for x in t["members"]))
    rival_team = next(
        i for i, t in enumerate(base["teams"])
        if any(x["position"] == "가드" and x["position_rank"] == 2 for x in t["members"])
    )
    moved = [[x["id"] for x in t["members"] if x["id"] != ace["id"]] for t in base["teams"]]
    moved[rival_team].append(ace["id"])
    res = client.post(
        "/api/teams/rearrange",
        json={"assignment": moved, "team_names": ["레드", "블루", "화이트"]},
    )
    assert res.status_code == 200, res.text
    after = res.json()
    assert after["manual"] is True, "수동 조정 표시가 없음"
    assert [t["name"] for t in after["teams"]] == ["레드", "블루", "화이트"]
    assert any(x["id"] == ace["id"] for x in after["teams"][rival_team]["members"]), "인원이 옮겨지지 않음"
    assert after["metrics"]["top_separation_violations"] >= 1, "옮겨서 생긴 제약 위반이 반영되지 않음"
    print(
        f"[manual] 가드 1위를 {ace_team + 1}팀 → {rival_team + 1}팀 이동: "
        f"편차 {base['metrics']['level_deviation_rate_avg']}% → {after['metrics']['level_deviation_rate_avg']}%, "
        f"역량분산위반 {base['metrics']['top_separation_violations']} → {after['metrics']['top_separation_violations']}"
    )
    assert client.get("/api/teams/result").json()["manual"] is True, "수동 조정 결과가 저장되지 않음"

    # 잘못된 요청 방어
    flat = [i for g in moved for i in g]
    assert client.post("/api/teams/rearrange", json={"assignment": [flat, [], []]}).status_code == 400
    assert client.post("/api/teams/rearrange", json={"assignment": [[flat[0]], [flat[0]], flat[2:]]}).status_code == 400
    assert client.post("/api/teams/rearrange", json={"assignment": [["없는id"], moved[1], moved[2]]}).status_code == 400
    print("[manual] 빈 팀 · 중복 배정 · 없는 인원 요청은 400 OK")

    # 8) 결과 조회 & xlsx 내보내기
    assert client.get("/api/teams/result").status_code == 200
    res = client.post("/api/teams/export", json={"team_names": ["레드", "블루", "화이트"]})
    assert res.status_code == 200 and len(res.content) > 3000
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(res.content))
    assert wb.sheetnames == ["팀편성결과", "팀요약"]
    assert [c.value for c in wb["팀편성결과"][1]][:8] == [
        "팀", "이름", "포지션", "역량평균", "나이보너스", "실질전력", "역량순위", "신장순위"]
    print(f"[export] xlsx {len(res.content)} bytes, sheets={wb.sheetnames}")

    # 9) 삭제 / 정리
    assert client.delete(f"/api/members/{fixed_id}").status_code == 204
    assert len(client.get("/api/members").json()) == 27
    assert client.delete("/api/members").status_code == 204
    assert client.get("/api/members").json() == []

    # 10) 인원 부족 시 오류
    res = client.post("/api/teams/generate", json={})
    assert res.status_code == 400
    print("[validate] 최소 인원 검증 OK")

    # 11) 정적 페이지
    for path in ("/", "/result", "/static/app.js", "/static/style.css"):
        assert client.get(path).status_code == 200, path
    print("[pages] 정적 파일 서빙 OK")

    print("\n전체 통과")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
